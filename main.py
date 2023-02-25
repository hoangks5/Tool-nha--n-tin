import openpyxl
import os
import random
import pandas as pd
import numpy as np
import json

from asyncio import events
from os import name
import re
from telethon import TelegramClient, events, hints, types
from telethon.tl.functions.channels import JoinChannelRequest
import time
from telethon.tl.types import InputPeerChat
from telethon.tl.functions.messages import SendMessageRequest
import random
from telethon.tl.types import messages
from telethon.tl.functions.account import UpdateUsernameRequest
from telethon.tl.functions.photos import UploadProfilePhotoRequest
from telethon.tl.functions.account import UpdateProfileRequest
import os
from datetime import datetime


def create_file(id_group):
    
    # Tạo một tệp Excel mới
    workbook = openpyxl.Workbook()

    # Lấy sheet đầu tiên của tệp Excel
    sheet = workbook.active

    # Thêm tiêu đề vào sheet
    sheet['A1'] = 'Thời gian'
    sheet['B1'] = 'Mã cộng'
    sheet['C1'] = 'Tỷ giá'
    sheet['D1'] = 'Thời gian'
    sheet['E1'] = 'Mã trừ'
    sheet['F1'] = 'Cần trả'
    sheet['G1'] = 'Còn lại'
    sheet['F2'] = 0
    sheet['G2'] = 0





    # Lưu tệp Excel
    workbook.save('data/'+str(id_group)+'.xlsx')

def convert_to_number(char):
    # Convert character to uppercase
    char = char.upper()
    # Get Unicode code point for character
    code = ord(char)
    # Subtract Unicode code point for 'A' to get 1-based index
    return code - ord('A') + 1

def them_ma_cong(id_group,id_column,code):
    workbook = openpyxl.load_workbook('data/'+str(id_group)+'.xlsx')

    # Select active worksheet
    worksheet = workbook.active

    last_row = 1
    for cell in worksheet[id_column]:
        if cell.value is not None:
            last_row += 1
    worksheet.cell(row=last_row, column=convert_to_number(id_column)).value = code

    # Save the changes
    workbook.save('data/'+str(id_group)+'.xlsx')

def tinh_toan(id_group):
    # Đọc dữ liệu từ bảng tính Excel
    df = pd.read_excel('data/'+str(id_group)+'.xlsx')

    
    column_b = df.iloc[:, 1].tolist()
    matrix_b = np.array(column_b).reshape((-1, 1))
    column_c = df.iloc[:, 2].tolist()
    matrix_c = np.array(column_c).reshape((-1, 1))
    column_d = df.iloc[:, 4].tolist()
    matrix_d = np.array(column_d).reshape((-1, 1))
    matrix_b = matrix_b[~np.isnan(matrix_b)].reshape(-1, 1)
    matrix_c = matrix_c[~np.isnan(matrix_c)].reshape(-1, 1)
    matrix_d = matrix_d[~np.isnan(matrix_d)].reshape(-1, 1)
    column_e = df.iloc[:, 0].tolist()
    matrix_e = np.array(column_e).reshape((-1, 1))
    column_f = df.iloc[:, 3].tolist()
    matrix_f = np.array(column_f).reshape((-1, 1))
    try:
        _tongnhap = sum(matrix_b*matrix_c)[0]
    except TypeError:
        _tongnhap = 0
    try:
        _trongtra = sum(matrix_d)[0]
    except TypeError:
        _trongtra = 0
    _conlai =  _tongnhap - _trongtra

    workbook = openpyxl.load_workbook('data/'+str(id_group)+'.xlsx')

    worksheet = workbook.active
    worksheet['F2'] = _tongnhap
    worksheet['G2'] = _conlai
    # Save the changes
    workbook.save('data/'+str(id_group)+'.xlsx')
    return {
        'time_cong' : matrix_e,
        'ma_cong': matrix_b,
        'ty_gia' : matrix_c,
        'time_tru' : matrix_f,
        'ma_tru' : matrix_d,
        'tong_nhap': _tongnhap,
        'tong_tra': _trongtra,
        'con_lai': _conlai
    }

def sua_ma_cong(id_group,stt,code):
    workbook = openpyxl.load_workbook('data/'+str(id_group)+'.xlsx')

    # Select active worksheet
    worksheet = workbook.active

    worksheet['B'+str(stt+1)] = code

    # Save the changes
    workbook.save('data/'+str(id_group)+'.xlsx')


def sua_ma_nhan(id_group,stt,code):
    workbook = openpyxl.load_workbook('data/'+str(id_group)+'.xlsx')

    # Select active worksheet
    worksheet = workbook.active

    worksheet['E'+str(stt+1)] = code

    # Save the changes
    workbook.save('data/'+str(id_group)+'.xlsx')

def update_admin(admin):
    with open('data/admin.txt','w',encoding='utf-8') as f:
        f.write('\n'.join(admin))
        f.close()

def update_group_id(id_group):
    str_id = []
    for id in id_group:
        str_id.append(str(id))
    with open('data/group_id.txt','w',encoding='utf-8') as f:
        f.write('\n'.join(str_id))
        f.close()


def info_admin_group():
    with open('data/admin.txt','r',encoding='utf-8') as f:
        accmin = f.read().splitlines()
    with open('data/group_id.txt','r',encoding='utf-8') as f:
        grup = f.read().splitlines()
    accminvc = []
    grupvc = []
    for i in accmin:
        accminvc.append('@'+i)
    for i in grup:
        grupvc.append('https://web.telegram.org/z/#-'+i)
    text = 'Danh sách admin:\n'+'\n'.join(accminvc)+'\n\nDanh sách group id:\n'+'\n'.join(grupvc)
    return text    


def info():
    code_chat = {
        'Bắt Đầu' : 'Reset vào tạo 1 phiên giao dịch mới',
        'vnd 2345' : 'Thiết lập tỷ giá VND = 2345',
        '+2000' : 'Thêm 1 mã nhập 2000',
        '-2000' : 'Thêm một lệnh trừ VND 2000',
        'sửa mã 1 = 2000' : 'Sửa mã nhập thứ 1 thành 2000',
        'sửa vnd 1 = 2000' : 'Sửa mã trừ thứ 1 thành 2000',
        'xuất file': 'Xuất ra file excel',
        'admin@hoangkss5' : 'Thêm @hoangkss5 làm admin',
        'deladmin@hoangkss5': 'Xoá @hoangkss5 khỏi danh sách admin',
        'id 122412414': 'Thêm id group 122412414 vào danh sách group làm việc',
        'delid 122412414' : 'Xoá id group 122412414 khỏi danh sách group làm việc'
    }
    mess = []
    for key in list(code_chat.keys()):
        mess.append('<b>'+key+'</b>  :  '+code_chat[key])
    return '\n'.join(mess)
api_id = 8759328 
api_hash = '1a270788cb618993f54f514f5a8c93c4'

ADMIN = []
ID_GROUP = []

with open('data/admin.txt','r',encoding='utf-8') as f:
    ADMIN = f.read().splitlines()
with open('data/group_id.txt','r',encoding='utf-8') as f:
    for id in f.read().splitlines():
        ID_GROUP.append(int(id))
try:
    with open('data/tygia.json') as f:
        ty_gia_vnd = json.load(f)
except:
    ty_gia_vnd = {}


def update_ty_gia(ty_gia):
    with open('data/tygia.json', 'w') as outfile:
        json.dump(ty_gia, outfile)


def messager_text(id_group):
    text = tinh_toan(id_group)
    cout = 0
    text_ma_cong = []
    for i in text['ma_cong']:
        text_ma_cong.append('<b>['+str(cout+1)+']</b> <i>'+str(text['time_cong'][cout][0]) +'</i> : <b>'+str(i[0]) + '</b> * ' + str(text['ty_gia'][cout][0])+' = ' + str("{:,}".format(int(i[0]*text['ty_gia'][cout][0])))       )
        cout += 1
    t1 = 'Tổng mã nhập: '+str(cout)+' mã\n' + '\n'.join(text_ma_cong)
    cout = 0
    text_ma_tru = []
    for i in text['ma_tru']:
        text_ma_tru.append('<b>['+str(cout+1)+']</b> <i>'+str(text['time_tru'][cout][0]) +'</i> : '+str("{:,}".format(int(i[0]))))
        cout += 1
    t2 = 'Tổng mã trừ: '+str(cout)+' mã\n' + '\n'.join(text_ma_tru)
    can_tra = int(text['tong_nhap'])
    con_lai = int(text['con_lai'])
    return t1+'\n\n\n'+t2+'\n\n\nCần trả: <b>'+str("{:,}".format(can_tra))+'</b>\nCòn lại: <b>'+str("{:,}".format(con_lai))+'</b>'

client = TelegramClient('admin', api_id, api_hash).start()
@client.on(events.NewMessage())
async def main(event):
    global ty_gia_vnd, ADMIN, ID_GROUP
    try:
        group = await event.get_chat()
    except:
        pass
    message = event.message
    sender = await message.get_sender()
    sender_username = sender.username
    if group.id in ID_GROUP and sender_username in ADMIN:

        if event.message.raw_text == 'Bắt Đầu':
            await client.send_message(group.id,'Bắt Đầu Ghi Giao Dịch Mới')
            create_file(group.id)


        if 'vnd' == event.message.raw_text.lower().split(' ')[0]:
            number = ''.join(re.findall(r'\d+\.\d+', event.message.raw_text))
            if number == '':
                number = ''.join(re.findall(r'\d', event.message.raw_text))
            await client.send_message(group.id,'Thiết lập tỷ giá VND '+number)
            ty_gia_vnd.update({str(group.id):int(number)})
            update_ty_gia(ty_gia=ty_gia_vnd)

        if 'id' == event.message.raw_text.lower().split(' ')[0]:
            number = ''.join(re.findall(r'\d+\.\d+', event.message.raw_text))
            if number == '':
                number = ''.join(re.findall(r'\d', event.message.raw_text))
            await client.send_message(group.id,'Thêm id gruop: '+number)
            ID_GROUP.append(int(number))
            update_group_id(ID_GROUP)
        if 'delid' == event.message.raw_text.lower().split(' ')[0]:
            number = ''.join(re.findall(r'\d+\.\d+', event.message.raw_text))
            if number == '':
                number = ''.join(re.findall(r'\d', event.message.raw_text))
            ID_GROUP.remove(int(number))
            await client.send_message(group.id,'Xoá id gruop: '+number)
            update_group_id(ID_GROUP)


        if 'admin' == event.message.raw_text.split('@')[0]:
            username = event.message.raw_text.split('@')[1]
            ADMIN.append(username)
            await client.send_message(group.id,'Thêm admin @'+username)
            update_admin(ADMIN)

        if 'deladmin' == event.message.raw_text.lower().split('@')[0]:
            username = event.message.raw_text.lower().split('@')[1]
            ADMIN.remove(username)
            await client.send_message(group.id,'Xoá admin @'+username)
            update_admin(ADMIN)

        
            
            





        if event.message.raw_text[0] == '+':
            now = datetime.now()
            current_time = now.strftime("%H:%M:%S")
            number = ''.join(re.findall(r'\d+\.\d+', event.message.raw_text))
            if number == '':
                number = ''.join(re.findall(r'\d', event.message.raw_text))

            them_ma_cong(group.id,'A',current_time)
            them_ma_cong(group.id,'B',number)
            them_ma_cong(group.id,'C',ty_gia_vnd[str(group.id)])
            await client.send_message(group.id,messager_text(group.id)+'\n\nTỷ giá VND: <b>'+str(ty_gia_vnd[str(group.id)])+'</b>',parse_mode='html')

            
        if event.message.raw_text[0] == '-':
            now = datetime.now()
            current_time = now.strftime("%H:%M:%S")
            number = ''.join(re.findall(r'\d+\.\d+', event.message.raw_text))
            if number == '':
                number = ''.join(re.findall(r'\d', event.message.raw_text))

            them_ma_cong(group.id,'D',current_time)
            them_ma_cong(group.id,'E',number)
            await client.send_message(group.id,messager_text(group.id)+'\n\nTỷ giá VND: <b>'+str(ty_gia_vnd[str(group.id)])+'</b>',parse_mode='html')

        if 'sửa mã' in event.message.raw_text.lower().split('=')[0]:
            number = ''.join(re.findall(r'\d', event.message.raw_text.lower().split('=')[0]))
            value = ''.join(re.findall(r'\d+\.\d+', event.message.raw_text.lower().split('=')[1]))
            if value == '':
                value = ''.join(re.findall(r'\d', event.message.raw_text.lower().split('=')[1]))
            sua_ma_cong(group.id,int(number),int(value))
            await client.send_message(group.id,messager_text(group.id)+'\n\nTỷ giá VND: <b>'+str(ty_gia_vnd[str(group.id)])+'</b>',parse_mode='html')


        if 'sửa vnd' in event.message.raw_text.lower().split('=')[0]:
            number = ''.join(re.findall(r'\d', event.message.raw_text.lower().split('=')[0]))
            value = ''.join(re.findall(r'\d+\.\d+', event.message.raw_text.lower().split('=')[1]))
            if value == '':
                value = ''.join(re.findall(r'\d', event.message.raw_text.lower().split('=')[1]))
            sua_ma_nhan(group.id,int(number),int(value))
            await client.send_message(group.id,messager_text(group.id)+'\n\nTỷ giá VND: <b>'+str(ty_gia_vnd[str(group.id)])+'</b>',parse_mode='html')



        if 'xuất file' == event.message.raw_text.lower():
            await client.send_file(group.id,'data/'+str(group.id)+'.xlsx')

        if 'hdsd' == event.message.raw_text.lower():
            await client.send_message(group.id,info(),parse_mode='html')
            await client.send_message(group.id,info_admin_group(),parse_mode='html')



client.run_until_disconnected()